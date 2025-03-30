import os
import pytest
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.dimensions import RowDimension, ColumnDimension
from openpyxl.worksheet.datavalidation import DataValidation
from mcp_excel import read_excel, get_excel_properties
from mcp_excel.main import main, mcp
import sys
from io import StringIO
from unittest.mock import patch, MagicMock
import importlib

# Test data directory
TEST_DATA_DIR = os.path.join(os.path.dirname(__file__), "test_data")

@pytest.fixture(scope="session")
def sample_excel_file():
    """Create a sample Excel file with various features for testing."""
    # Create test directory if it doesn't exist
    os.makedirs(TEST_DATA_DIR, exist_ok=True)
    
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Add some data
    ws['A1'] = 'Name'
    ws['B1'] = 'Age'
    ws['A2'] = 'John'
    ws['B2'] = 25
    ws['A3'] = 'Jane'
    ws['B3'] = 30
    
    # Add data validation (dropdown list)
    dv = DataValidation(
        type="list",
        formula1='"Young,Adult,Senior"',
        allow_blank=True,
        showErrorMessage=True,
        showInputMessage=True
    )
    ws.add_data_validation(dv)
    dv.add('C2:C3')
    
    # Add data validation (number range)
    dv2 = DataValidation(
        type="whole",
        operator="between",
        formula1="0",
        formula2="100",
        allow_blank=True,
        showErrorMessage=True,
        showInputMessage=True
    )
    ws.add_data_validation(dv2)
    dv2.add('D2:D3')
    
    # Merge cells
    ws.merge_cells('E1:F1')
    ws['E1'] = 'Merged Header'
    
    # Hide a row and column
    # Create row dimension with hidden=True
    rd = RowDimension(ws, index=4, hidden=True)
    ws.row_dimensions[4] = rd
    
    # Create column dimension with hidden=True
    cd = ColumnDimension(ws, index='G', hidden=True)
    ws.column_dimensions['G'] = cd
    
    # Save the file
    file_path = os.path.join(TEST_DATA_DIR, "test.xlsx")
    wb.save(file_path)
    
    yield file_path
    
    # Cleanup after all tests are done
    try:
        if os.path.exists(TEST_DATA_DIR):
            for file in os.listdir(TEST_DATA_DIR):
                os.remove(os.path.join(TEST_DATA_DIR, file))
            os.rmdir(TEST_DATA_DIR)
    except:
        pass

def test_read_excel_basic(sample_excel_file):
    """Test basic Excel reading functionality."""
    df, properties = read_excel(sample_excel_file)
    
    # Check DataFrame content
    assert isinstance(df, pd.DataFrame)
    assert len(df) == 2  # 2 data rows (excluding header)
    assert list(df.columns) == ['Name', 'Age', 'Unnamed: 2', 'Unnamed: 3', 'Merged Header']
    assert df.iloc[0]['Name'] == 'John'
    assert df.iloc[0]['Age'] == 25
    
    # Check properties
    assert isinstance(properties, dict)
    assert 'data_validation' in properties
    assert 'dropdown_lists' in properties
    assert 'merged_cells' in properties
    assert 'hidden_rows' in properties
    assert 'hidden_columns' in properties

def test_read_excel_with_validation(sample_excel_file):
    """Test reading Excel with data validation."""
    df, properties = read_excel(sample_excel_file)
    
    # Check data validation
    assert len(properties['data_validation']) == 2  # Two validation rules
    
    # Check dropdown list
    dropdowns = [dv for dv in properties['data_validation'] if dv['type'] == 'list']
    assert len(dropdowns) == 1
    assert dropdowns[0]['cell'] == 'C2:C3'
    assert dropdowns[0]['formula1'] == '"Young,Adult,Senior"'
    
    # Check number range validation
    number_validations = [dv for dv in properties['data_validation'] if dv['type'] == 'whole']
    assert len(number_validations) == 1
    assert number_validations[0]['cell'] == 'D2:D3'
    assert number_validations[0]['operator'] == 'between'
    assert number_validations[0]['formula1'] == '0'
    assert number_validations[0]['formula2'] == '100'

def test_read_excel_with_merged_cells(sample_excel_file):
    """Test reading Excel with merged cells."""
    df, properties = read_excel(sample_excel_file)
    
    # Check merged cells
    assert len(properties['merged_cells']) == 1
    assert str(properties['merged_cells'][0]) == 'E1:F1'

def test_read_excel_with_hidden(sample_excel_file):
    """Test reading Excel with hidden rows and columns."""
    df, properties = read_excel(sample_excel_file)
    
    # Check hidden rows and columns
    assert 4 in properties['hidden_rows']
    assert 7 in properties['hidden_columns']  # G column

def test_read_excel_errors():
    """Test error handling in Excel reading."""
    # Test non-existent file
    with pytest.raises(FileNotFoundError):
        read_excel("nonexistent.xlsx")
    
    # Test non-existent sheet
    with pytest.raises(ValueError):
        read_excel(os.path.join(TEST_DATA_DIR, "test.xlsx"), sheet_name="NonexistentSheet")
    
    # Test general exception
    with pytest.raises(Exception):
        read_excel(None)

def test_get_excel_properties(sample_excel_file):
    """Test getting Excel properties."""
    properties = get_excel_properties(sample_excel_file)
    
    # Check properties structure
    assert isinstance(properties, dict)
    assert 'data_validation' in properties
    assert 'dropdown_lists' in properties
    assert 'merged_cells' in properties
    assert 'hidden_rows' in properties
    assert 'hidden_columns' in properties
    
    # Check specific properties
    assert len(properties['data_validation']) == 2
    assert len(properties['merged_cells']) == 1
    assert 4 in properties['hidden_rows']
    assert 7 in properties['hidden_columns']

def test_get_excel_properties_errors():
    """Test error handling in get_excel_properties."""
    # Test non-existent file
    with pytest.raises(FileNotFoundError):
        get_excel_properties("nonexistent.xlsx")
    
    # Test non-existent sheet
    with pytest.raises(ValueError):
        get_excel_properties(os.path.join(TEST_DATA_DIR, "test.xlsx"), sheet_name="NonexistentSheet")
    
    # Test general exception
    with pytest.raises(Exception):
        get_excel_properties(None)

def test_main_function(capsys):
    """Test the main function."""
    # Mock the MCP server's run method
    with patch.object(mcp, 'run') as mock_run:
        # Call main()
        main()
        
        # Check that mcp.run() was called
        mock_run.assert_called_once()
        
        # Check that the startup message was printed
        captured = capsys.readouterr()
        assert "Starting MCP server for Excel operations..." in captured.err

def test_read_excel_with_sheet_name(sample_excel_file, capsys):
    """Test reading Excel with specific sheet name."""
    # Test with sheet name
    df, properties = read_excel(sample_excel_file, sheet_name="Sheet1")
    
    # Check that the correct message was printed
    captured = capsys.readouterr()
    assert f"Reading sheet 'Sheet1' from {sample_excel_file}" in captured.out
    
    # Test without sheet name
    df, properties = read_excel(sample_excel_file)
    
    # Check that the correct message was printed
    captured = capsys.readouterr()
    assert f"No specific sheet requested. Reading the first sheet from {sample_excel_file}" in captured.out

def test_main_module_execution():
    """Test that the main function is called when the module is run directly."""
    with patch('mcp_excel.main.mcp') as mock_mcp:
        # Mock the mcp.run() method
        mock_mcp.run.return_value = None
        
        # Save the original __name__
        original_name = mcp_excel.main.__name__
        
        try:
            # Set __name__ to '__main__'
            mcp_excel.main.__name__ = '__main__'
            
            # Call main() directly
            mcp_excel.main.main()
            
            # Verify that mcp.run() was called
            mock_mcp.run.assert_called_once()
        finally:
            # Restore the original __name__
            mcp_excel.main.__name__ = original_name 