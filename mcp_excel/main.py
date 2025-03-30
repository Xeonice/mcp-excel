import pandas as pd
from mcp.server.fastmcp import FastMCP
from openpyxl import load_workbook
from typing import Dict, Any, Optional, Tuple
import sys

mcp = FastMCP(
    "excel_access",
    dependencies=["pandas", "openpyxl"],
    description="MCP server for Excel file operations"
)

def _get_hidden_dimensions(ws) -> Tuple[list, list]:
    """
    Get hidden rows and columns from a worksheet.
    
    Args:
        ws: The worksheet to analyze.
        
    Returns:
        Tuple[list, list]: A tuple containing:
            - List of hidden row numbers
            - List of hidden column numbers
    """
    hidden_rows = []
    hidden_columns = []
    
    # Get hidden rows
    for row_idx, rd in ws.row_dimensions.items():
        if rd.hidden:
            hidden_rows.append(row_idx)
    
    # Get hidden columns
    for col_letter, cd in ws.column_dimensions.items():
        if cd.hidden:
            # Convert column letter to number (A=1, B=2, etc.)
            col_idx = 0
            for i, c in enumerate(reversed(col_letter)):
                col_idx += (ord(c) - ord('A') + 1) * (26 ** i)
            hidden_columns.append(col_idx)
    
    return hidden_rows, hidden_columns

@mcp.tool()
def read_excel(file_path: str, sheet_name: str = None) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    """
    Read an Excel file and return its content as a pandas DataFrame along with its properties.
    
    Args:
        file_path (str): Path to the Excel file.
        sheet_name (str, optional): Name or index of the sheet to read. 
                                   If None, reads the first sheet by default.
    
    Returns:
        Tuple[pd.DataFrame, Dict[str, Any]]: A tuple containing:
            - DataFrame containing the Excel sheet data
            - Dictionary containing sheet properties including:
                - data_validation: List of cells with data validation
                - dropdown_lists: List of cells with dropdown lists
                - merged_cells: List of merged cell ranges
                - hidden_rows: List of hidden row numbers
                - hidden_columns: List of hidden column numbers
        
    Raises:
        FileNotFoundError: If the specified file does not exist.
        ValueError: If the specified sheet does not exist in the Excel file.
    """
    try:
        # Read the Excel file for properties
        wb = load_workbook(file_path, data_only=True)
        try:
            ws = wb[sheet_name] if sheet_name else wb.active
        except KeyError:
            raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file.")
        
        # Get hidden dimensions
        hidden_rows, hidden_columns = _get_hidden_dimensions(ws)
        
        # Get properties
        properties = {
            "data_validation": [],
            "dropdown_lists": [],
            "merged_cells": list(ws.merged_cells.ranges) if ws.merged_cells else [],
            "hidden_rows": hidden_rows,
            "hidden_columns": hidden_columns
        }
        
        # Get data validation and dropdown lists
        for dv in ws.data_validations.dataValidation:
            # Get the range of cells this validation applies to
            cell_range = dv.sqref  # Use sqref instead of cells
            validation_info = {
                "cell": cell_range,
                "type": dv.type,
                "operator": dv.operator,
                "formula1": dv.formula1,
                "formula2": dv.formula2,
                "allow_blank": dv.allow_blank,
                "show_error": dv.showErrorMessage,
                "show_input": dv.showInputMessage
            }
            properties["data_validation"].append(validation_info)
            
            # Check if it's a dropdown list
            if dv.type == "list":
                # Clean up the formula to get the list options
                options = dv.formula1.strip('"').split(',')
                properties["dropdown_lists"].append({
                    "cell": cell_range,
                    "options": options
                })
        
        # Read the Excel file content
        if sheet_name is None:
            print(f"No specific sheet requested. Reading the first sheet from {file_path}")
            df = pd.read_excel(file_path, engine="openpyxl")
        else:
            print(f"Reading sheet '{sheet_name}' from {file_path}")
            df = pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl")
                
        return df, properties
        
    except FileNotFoundError:
        raise FileNotFoundError(f"Excel file not found at path: {file_path}")
    except ValueError as e:
        raise e
    except Exception as e:
        raise Exception(f"Error reading Excel file: {str(e)}")

@mcp.tool()
def get_excel_properties(file_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """
    Get Excel file properties including data validation and dropdown lists.
    
    Args:
        file_path (str): Path to the Excel file.
        sheet_name (str, optional): Name of the sheet to analyze. If None, analyzes the first sheet.
    
    Returns:
        Dict[str, Any]: Dictionary containing sheet properties including:
            - data_validation: List of cells with data validation
            - dropdown_lists: List of cells with dropdown lists
            - merged_cells: List of merged cell ranges
            - hidden_rows: List of hidden row numbers
            - hidden_columns: List of hidden column numbers
    """
    try:
        wb = load_workbook(file_path, data_only=True)
        try:
            ws = wb[sheet_name] if sheet_name else wb.active
        except KeyError:
            raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file.")
        
        # Get hidden dimensions
        hidden_rows, hidden_columns = _get_hidden_dimensions(ws)
        
        properties = {
            "data_validation": [],
            "dropdown_lists": [],
            "merged_cells": list(ws.merged_cells.ranges) if ws.merged_cells else [],
            "hidden_rows": hidden_rows,
            "hidden_columns": hidden_columns
        }
        
        # Get data validation and dropdown lists
        for dv in ws.data_validations.dataValidation:
            # Get the range of cells this validation applies to
            cell_range = dv.sqref  # Use sqref instead of cells
            validation_info = {
                "cell": cell_range,
                "type": dv.type,
                "operator": dv.operator,
                "formula1": dv.formula1,
                "formula2": dv.formula2,
                "allow_blank": dv.allow_blank,
                "show_error": dv.showErrorMessage,
                "show_input": dv.showInputMessage
            }
            properties["data_validation"].append(validation_info)
            
            # Check if it's a dropdown list
            if dv.type == "list":
                # Clean up the formula to get the list options
                options = dv.formula1.strip('"').split(',')
                properties["dropdown_lists"].append({
                    "cell": cell_range,
                    "options": options
                })
        
        return properties
        
    except FileNotFoundError:
        raise FileNotFoundError(f"Excel file not found at path: {file_path}")
    except ValueError as e:
        raise e
    except Exception as e:
        raise Exception(f"Error reading Excel properties: {str(e)}")

def main():
    """Entry point for the MCP server."""
    print("Starting MCP server for Excel operations...", file=sys.stderr)
    mcp.run()

if __name__ == "__main__":
    main()